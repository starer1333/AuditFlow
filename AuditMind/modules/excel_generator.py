import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
from config import OUTPUT_DIR

def generate(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "银行存款余额调节表"
    ws.merge_cells("A1:F1")
    ws["A1"] = "银行存款余额调节表"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    bank = data.get("bank_name", "未识别")
    acc = data.get("account_number", "未识别")
    bal = data.get("ending_balance", 0)
    period = data.get("period", "未识别")
    conf = data.get("validation", {}).get("final_confidence", 0)
    row = 3
    for r in [["被审计单位","XX科技","","索引号","A-2-1"],
              ["银行名称",bank,"","账号",acc],
              ["对账单余额",f"{bal:,.2f}" if bal else "未识别","","期间",period],
              ["置信度",f"{conf*100:.0f}%","","需复核","是" if data.get("validation",{}).get("need_review") else "否"]]:
        for c,v in enumerate(r,1): ws.cell(row, c, v)
        row += 1
    row += 1
    for c,h in enumerate(["项目","金额","审计标识","说明"],1):
        cell = ws.cell(row, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D3D3D3")
    row += 1
    for item, amt, mark, note in [["银行对账单余额",bal,"B","系统识别"],
                                   ["加：企业已收银行未收","","",""],
                                   ["减：企业已付银行未付","","",""],
                                   ["调节后余额",bal,"G",""],
                                   ["企业账面余额","","","待填写"],
                                   ["差异","","",""]]:
        ws.cell(row, 1, item)
        if amt: ws.cell(row, 2, amt).number_format = '#,##0.00'
        ws.cell(row, 3, mark)
        ws.cell(row, 4, note)
        row += 1
    row += 1
    ws.cell(row, 1, f"审计意见: {data.get('risk_notes', '无异常')}")
    row += 2
    ws.cell(row, 1, f"编制人: AuditFlow  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    for i,w in enumerate([20,18,12,35],1): ws.column_dimensions[chr(64+i)].width = w
    p = os.path.join(OUTPUT_DIR, f"银行余额调节表_{bank}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(p)
    return p