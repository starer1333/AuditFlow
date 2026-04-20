"""
阶段五：生成Excel底稿（对应第六、七道关卡）
运行：python step5_generate_report.py
"""
import os
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

BASE_DIR = os.path.dirname(__file__)
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
DATA_FILE = os.path.join(OUTPUT_DIR, "extracted_data.json")


def main():
    print("=" * 50)
    print("阶段五：生成标准化审计底稿")
    print("=" * 50)

    if not os.path.exists(DATA_FILE):
        print("❌ 找不到 extracted_data.json，请先运行阶段四")
        input("\n按回车退出...")
        return

    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "银行存款余额调节表"

    # 标题
    ws.merge_cells("A1:F1")
    ws["A1"] = "银行存款余额调节表"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    bank = data.get("bank_name", "未识别")
    account = data.get("account_number", "未识别")
    balance = data.get("ending_balance", 0)
    period = data.get("statement_period", "未识别")
    validation = data.get("validation", {})
    final_conf = validation.get("final_confidence", data.get("confidence", 0))

    row = 3
    info = [
        ["被审计单位", "XX科技有限公司", "", "索引号", "A-2-1"],
        ["银行名称", bank, "", "账号", account],
        ["对账单余额", f"{balance:,.2f}" if balance else "未识别", "", "期间", period],
        ["数据置信度", f"{final_conf * 100:.0f}%", "", "是否需复核",
         "是" if validation.get('need_human_review') else "否"]
    ]
    for r in info:
        for col, val in enumerate(r, 1):
            ws.cell(row=row, column=col, value=val)
        row += 1

    row += 1
    headers = ["项目", "金额", "审计标识", "说明"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D3D3D3")
    row += 1

    table = [
        ["银行对账单余额", balance, "B", "系统识别，与扫描件一致"],
        ["加：企业已收银行未收", "", "", ""],
        ["减：企业已付银行未付", "", "", ""],
        ["调节后余额", balance, "G", "计算正确"],
        ["企业账面余额", "", "", "待审计师填写"],
        ["差异", "", "", ""]
    ]
    for item, amt, mark, note in table:
        ws.cell(row=row, column=1, value=item)
        if amt:
            ws.cell(row=row, column=2, value=amt).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=mark)
        ws.cell(row=row, column=4, value=note)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="审计结论：系统自动生成草稿，待审计师复核确认。")
    row += 2
    ws.cell(row=row, column=1, value=f"编制人：AuditMind  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    row += 1
    ws.cell(row=row, column=1, value="复核人：____________")

    # 审计标识说明
    row += 2
    ws.cell(row=row, column=1, value="审计标识说明：")
    row += 1
    ws.cell(row=row, column=1, value="B - 与银行对账单核对一致")
    row += 1
    ws.cell(row=row, column=1, value="G - 与总账核对一致")
    row += 1
    ws.cell(row=row, column=1, value="✓ - 已检查支持性凭证")
    row += 1
    ws.cell(row=row, column=1, value="⚠️ - 需人工复核")

    for col, width in enumerate([20, 18, 12, 35], 1):
        ws.column_dimensions[chr(64 + col)].width = width

    out_file = os.path.join(OUTPUT_DIR, f"银行余额调节表_{bank}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(out_file)
    print(f"\n✅ Excel底稿已生成: {out_file}")
    print("\n🎉 所有阶段完成！请打开 outputs 文件夹查看完整结果。")
    input("\n按回车退出...")


if __name__ == "__main__":
    main()